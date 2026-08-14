from __future__ import annotations

import base64
import io
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from .models import (
    Absence,
    AccountantOperation,
    ActionLog,
    AttendanceEvent,
    Debt,
    Employee,
    ExpenseCategory,
    MoneyMovement,
    MoneyOperation,
    Shift,
    User,
)
from .serializers import cash_income_total, expected_cash_end, expense_tree_dto
from .timeutil import app_tz, now_local


HEADER_FILL = PatternFill("solid", fgColor="1B1BB3")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1B1BB3")
MONEY = "0.00"


def _parse_date(raw: str | None, fallback: date) -> date:
    if not raw:
        return fallback
    return date.fromisoformat(str(raw).strip()[:10])


def default_report_range() -> tuple[date, date]:
    today = now_local().date()
    start = today - timedelta(days=30)
    return start, today


def _style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws) -> None:
    for col in ws.columns:
        width = 12
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                width = max(width, min(48, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _money(value: Decimal | float | int | None) -> float:
    if value is None:
        return 0.0
    return float(Decimal(str(value)).quantize(Decimal("0.01")))


def _wb_to_payload(wb: Workbook, filename: str) -> dict[str, Any]:
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    return {
        "filename": filename,
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "content_base64": base64.b64encode(raw).decode("ascii"),
        "size": len(raw),
    }


async def _shifts_in_range(session: AsyncSession, d_from: date, d_to: date) -> list[Shift]:
    return list(
        (
            await session.execute(
                select(Shift)
                .options(
                    selectinload(Shift.operations)
                    .selectinload(MoneyOperation.movements)
                    .selectinload(MoneyMovement.vault)
                )
                .where(Shift.shift_date >= d_from, Shift.shift_date <= d_to)
                .order_by(Shift.shift_date.asc())
            )
        )
        .scalars()
        .all()
    )


async def generate_report(
    session: AsyncSession,
    user: User,
    report_type: str,
    date_from: str | None = None,
    date_to: str | None = None,
) -> dict[str, Any]:
    if user.role != "superuser":
        raise PermissionError("Отчёты доступны только суперпользователю")

    default_from, default_to = default_report_range()
    d_from = _parse_date(date_from, default_from)
    d_to = _parse_date(date_to, default_to)
    if d_from > d_to:
        raise ValueError("Дата «С» не может быть позже даты «По»")

    kind = str(report_type or "").strip().lower()
    if kind in ("balance", "баланс"):
        return await _report_balance(session, d_from, d_to)
    if kind in ("odds", "оддс"):
        return await _report_odds(session, d_from, d_to)
    if kind in ("pl", "opiу", "opiu", "опиу", "pnl"):
        return await _report_pl(session, d_from, d_to)
    if kind in ("expenses", "расходы", "expense"):
        return await _report_expenses(session, d_from, d_to)
    if kind in ("attendance", "посещаемость"):
        return await _report_attendance(session, d_from, d_to)
    if kind in ("logs", "log"):
        return await _report_logs(session, d_from, d_to)
    raise ValueError("Неизвестный тип отчёта")


def _sheet_title(name: str, used: set[str]) -> str:
    cleaned = "".join(ch for ch in name if ch not in r"\/*?:[]").strip() or "Лист"
    base = cleaned[:31]
    title = base
    i = 2
    while title in used:
        suffix = f"_{i}"
        title = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(title)
    return title


def _minutes_between(start, end) -> int:
    s = start.hour * 60 + start.minute
    e = end.hour * 60 + end.minute
    if e < s:
        e += 24 * 60
    return max(0, e - s)


def _format_duration(minutes: int) -> str:
    h, m = divmod(max(0, int(minutes)), 60)
    return f"{h} час. {m:02d} мин."


ABSENCE_LABELS = {
    "day_off": "ВЫХОДНОЙ",
    "sick": "БОЛЬНИЧНЫЙ",
    "vacation": "ОТПУСК",
}


async def _report_balance(session: AsyncSession, d_from: date, d_to: date) -> dict[str, Any]:
    shifts = await _shifts_in_range(session, d_from, d_to)
    debts = list(
        (
            await session.execute(
                select(Debt)
                .where(Debt.created_at <= datetime.combine(d_to, datetime.max.time(), tzinfo=app_tz()))
                .order_by(Debt.id.asc())
            )
        )
        .scalars()
        .all()
    )
    # Outstanding as of period end: unsettled debts created on/before d_to
    receivable = sum((d.amount for d in debts if d.kind == "receivable" and not d.is_settled), Decimal("0"))
    payable = sum((d.amount for d in debts if d.kind == "payable" and not d.is_settled), Decimal("0"))

    cash_end = None
    cash_start_period = None
    if shifts:
        cash_start_period = shifts[0].cash_start
        for s in reversed(shifts):
            if s.cash_end is not None:
                cash_end = s.cash_end
                break
            if s.cash_start is not None and cash_end is None:
                cash_end = s.cash_start

    assets = _money(cash_end) + _money(receivable)
    liabilities = _money(payable)
    equity = assets - liabilities

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "БАЛАНС — Автосервис CARTA"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Период: {d_from.strftime('%d.%m.%Y')} — {d_to.strftime('%d.%m.%Y')}"
    ws.append([])
    ws.append(["Показатель", "Сумма, руб."])
    _style_header(ws, 4, 2)
    rows = [
        ("Касса на начало периода", _money(cash_start_period)),
        ("Касса на конец периода", _money(cash_end)),
        ("Дебиторская задолженность", _money(receivable)),
        ("Кредиторская задолженность", _money(payable)),
        ("Активы (касса + дебиторка)", assets),
        ("Обязательства (кредиторка)", liabilities),
        ("Чистые активы", equity),
    ]
    for name, value in rows:
        ws.append([name, value])
        ws.cell(row=ws.max_row, column=2).number_format = MONEY
    _autosize(ws)

    ws2 = wb.create_sheet("Касса по дням")
    ws2.append(["Дата", "Касса начало", "Касса конец", "Движения по кассе", "Ожидаемый конец", "Расхождение"])
    _style_header(ws2, 1, 6)
    for s in shifts:
        movements = cash_income_total(s)
        expected = expected_cash_end(s)
        mismatch = bool(getattr(s, "cash_mismatch", False))
        if s.cash_start is not None and s.cash_end is not None and expected is not None:
            mismatch = mismatch or (
                Decimal(str(s.cash_end)).quantize(Decimal("0.01"))
                != expected.quantize(Decimal("0.01"))
            )
        ws2.append(
            [
                s.shift_date.strftime("%d.%m.%Y"),
                _money(s.cash_start) if s.cash_start is not None else None,
                _money(s.cash_end) if s.cash_end is not None else None,
                _money(movements),
                _money(expected) if expected is not None else None,
                "Да" if mismatch else "Нет",
            ]
        )
        for col in range(2, 6):
            if ws2.cell(row=ws2.max_row, column=col).value is not None:
                ws2.cell(row=ws2.max_row, column=col).number_format = MONEY
    _autosize(ws2)

    ws3 = wb.create_sheet("Задолженности")
    ws3.append(["Дата", "Вид", "Направление", "Категория", "Название", "Контрагент", "Сумма", "Погашено"])
    _style_header(ws3, 1, 8)
    for d in debts:
        created = d.created_at.astimezone(app_tz()).date() if d.created_at else None
        if created and created > d_to:
            continue
        ws3.append(
            [
                created.strftime("%d.%m.%Y") if created else "",
                "Дебиторская" if d.kind == "receivable" else "Кредиторская",
                "Приход" if d.direction == "income" else "Расход",
                d.category,
                d.title,
                d.counterparty or "",
                _money(d.amount),
                "Да" if d.is_settled else "Нет",
            ]
        )
        ws3.cell(row=ws3.max_row, column=7).number_format = MONEY
    _autosize(ws3)

    filename = f"CARTA_BALANS_{d_from.strftime('%Y%m%d')}_{d_to.strftime('%Y%m%d')}.xlsx"
    return _wb_to_payload(wb, filename)


async def _report_odds(session: AsyncSession, d_from: date, d_to: date) -> dict[str, Any]:
    shifts = await _shifts_in_range(session, d_from, d_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "ОДДС"
    ws["A1"] = "ОДДС — расхождения кассы"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Период: {d_from.strftime('%d.%m.%Y')} — {d_to.strftime('%d.%m.%Y')}"
    ws.append([])
    ws.append(
        [
            "Дата смены",
            "Касса начало",
            "Касса конец",
            "Ожидаемый конец",
            "Разница",
            "Движения по кассе",
        ]
    )
    _style_header(ws, 4, 6)

    count = 0
    total_diff = Decimal("0")
    for s in shifts:
        expected = expected_cash_end(s)
        if s.cash_start is None or s.cash_end is None or expected is None:
            continue
        entered = Decimal(str(s.cash_end)).quantize(Decimal("0.01"))
        expected_q = expected.quantize(Decimal("0.01"))
        if entered == expected_q and not getattr(s, "cash_mismatch", False):
            continue
        diff = entered - expected_q
        count += 1
        total_diff += diff
        ws.append(
            [
                s.shift_date.strftime("%d.%m.%Y"),
                _money(s.cash_start),
                _money(entered),
                _money(expected_q),
                _money(diff),
                _money(cash_income_total(s)),
            ]
        )
        for col in range(2, 7):
            ws.cell(row=ws.max_row, column=col).number_format = MONEY

    ws.append([])
    ws.append(["Всего расхождений", count])
    ws.append(["Суммарная разница", _money(total_diff)])
    ws.cell(row=ws.max_row, column=2).number_format = MONEY
    _autosize(ws)

    filename = f"CARTA_ODDS_{d_from.strftime('%Y%m%d')}_{d_to.strftime('%Y%m%d')}.xlsx"
    return _wb_to_payload(wb, filename)


async def _report_pl(session: AsyncSession, d_from: date, d_to: date) -> dict[str, Any]:
    shifts = await _shifts_in_range(session, d_from, d_to)
    start_dt = datetime.combine(d_from, datetime.min.time(), tzinfo=app_tz())
    end_dt = datetime.combine(d_to, datetime.max.time(), tzinfo=app_tz())

    accountant_ops = list(
        (
            await session.execute(
                select(AccountantOperation)
                .where(
                    AccountantOperation.created_at >= start_dt,
                    AccountantOperation.created_at <= end_dt,
                )
                .order_by(AccountantOperation.created_at.asc())
            )
        )
        .scalars()
        .all()
    )
    debts = list(
        (
            await session.execute(
                select(Debt)
                .where(Debt.created_at >= start_dt, Debt.created_at <= end_dt)
                .order_by(Debt.created_at.asc())
            )
        )
        .scalars()
        .all()
    )

    income_rows: list[list[Any]] = []
    expense_rows: list[list[Any]] = []
    income_by_cat: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    expense_by_cat: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

    def add_row(
        bucket: list[list[Any]],
        totals: dict[str, Decimal],
        when: date,
        method: str,
        source: str,
        category: str,
        title: str,
        amount: Decimal,
        comment: str = "",
    ) -> None:
        bucket.append(
            [
                when.strftime("%d.%m.%Y"),
                method,
                source,
                category,
                title,
                _money(amount),
                comment,
            ]
        )
        totals[category] += Decimal(str(amount))

    for s in shifts:
        for op in s.operations:
            amount = sum((m.amount for m in op.movements), Decimal("0"))
            if amount <= 0:
                continue
            if op.direction == "income":
                add_row(
                    income_rows,
                    income_by_cat,
                    s.shift_date,
                    "Кассовый",
                    "Смена",
                    op.category,
                    op.title,
                    amount,
                    op.comment or "",
                )
            else:
                add_row(
                    expense_rows,
                    expense_by_cat,
                    s.shift_date,
                    "Кассовый",
                    "Смена",
                    op.category,
                    op.title,
                    amount,
                    op.comment or "",
                )

    for op in accountant_ops:
        when = op.created_at.astimezone(app_tz()).date() if op.created_at else d_from
        if op.direction == "income":
            add_row(
                income_rows,
                income_by_cat,
                when,
                "Кассовый",
                "Бухгалтерия",
                op.category,
                op.title,
                op.amount,
                op.comment or "",
            )
        else:
            add_row(
                expense_rows,
                expense_by_cat,
                when,
                "Кассовый",
                "Бухгалтерия",
                op.category,
                op.title,
                op.amount,
                op.comment or "",
            )

    for d in debts:
        when = d.created_at.astimezone(app_tz()).date() if d.created_at else d_from
        if d.direction == "income":
            add_row(
                income_rows,
                income_by_cat,
                when,
                "Начисление",
                "Задолженность",
                d.category,
                d.title,
                d.amount,
                d.comment or "",
            )
        else:
            add_row(
                expense_rows,
                expense_by_cat,
                when,
                "Начисление",
                "Задолженность",
                d.category,
                d.title,
                d.amount,
                d.comment or "",
            )

    total_income = sum(income_by_cat.values(), Decimal("0"))
    total_expense = sum(expense_by_cat.values(), Decimal("0"))
    profit = total_income - total_expense

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "ОПиУ — отчёт о прибылях и убытках"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Период: {d_from.strftime('%d.%m.%Y')} — {d_to.strftime('%d.%m.%Y')}"
    ws.append([])
    ws.append(["Показатель", "Сумма, руб."])
    _style_header(ws, 4, 2)
    for name, value in (
        ("Доходы всего", total_income),
        ("Расходы всего", total_expense),
        ("Прибыль / убыток", profit),
    ):
        ws.append([name, _money(value)])
        ws.cell(row=ws.max_row, column=2).number_format = MONEY
    ws.append([])
    ws.append(["Доходы по видам", "Сумма, руб."])
    _style_header(ws, ws.max_row, 2)
    for cat, value in sorted(income_by_cat.items(), key=lambda x: x[0]):
        ws.append([cat, _money(value)])
        ws.cell(row=ws.max_row, column=2).number_format = MONEY
    ws.append([])
    ws.append(["Расходы по видам", "Сумма, руб."])
    _style_header(ws, ws.max_row, 2)
    for cat, value in sorted(expense_by_cat.items(), key=lambda x: x[0]):
        ws.append([cat, _money(value)])
        ws.cell(row=ws.max_row, column=2).number_format = MONEY
    _autosize(ws)

    headers = ["Дата", "Метод", "Источник", "Вид", "Название", "Сумма", "Комментарий"]
    for title, rows in (("Доходы", income_rows), ("Расходы", expense_rows)):
        sheet = wb.create_sheet(title)
        sheet.append(headers)
        _style_header(sheet, 1, len(headers))
        for row in rows:
            sheet.append(row)
            sheet.cell(row=sheet.max_row, column=6).number_format = MONEY
        _autosize(sheet)

    filename = f"CARTA_OPiU_{d_from.strftime('%Y%m%d')}_{d_to.strftime('%Y%m%d')}.xlsx"
    return _wb_to_payload(wb, filename)


async def _report_expenses(session: AsyncSession, d_from: date, d_to: date) -> dict[str, Any]:
    cats = list(
        (
            await session.execute(
                select(ExpenseCategory)
                .options(selectinload(ExpenseCategory.items))
                .order_by(ExpenseCategory.sort_order)
            )
        )
        .scalars()
        .all()
    )
    tree = expense_tree_dto(cats)
    shifts = await _shifts_in_range(session, d_from, d_to)

    # amounts[(category, title)] from shift expenses + accountant cash expenses
    amounts: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    for s in shifts:
        for op in s.operations:
            if op.direction != "expense":
                continue
            total = sum((m.amount for m in op.movements), Decimal("0"))
            amounts[(op.category, op.title)] += total

    start_dt = datetime.combine(d_from, datetime.min.time(), tzinfo=app_tz())
    end_dt = datetime.combine(d_to, datetime.max.time(), tzinfo=app_tz())
    for op in (
        await session.execute(
            select(AccountantOperation).where(
                AccountantOperation.direction == "expense",
                AccountantOperation.created_at >= start_dt,
                AccountantOperation.created_at <= end_dt,
            )
        )
    ).scalars():
        amounts[(op.category, op.title)] += Decimal(str(op.amount))

    group_totals: dict[str, Decimal] = {}
    for group in tree:
        gname = group["name"]
        total = Decimal("0")
        for item in group["items"]:
            total += amounts.get((gname, item), Decimal("0"))
        group_totals[gname] = total

    wb = Workbook()
    used_titles: set[str] = set()
    ws = wb.active
    ws.title = _sheet_title("ИТОГО", used_titles)
    ws.append(["Группа расходов", "Сумма"])
    _style_header(ws, 1, 2)
    for group in tree:
        ws.append([group["name"], _money(group_totals.get(group["name"], Decimal("0")))])
        ws.cell(row=ws.max_row, column=2).number_format = MONEY
    grand = sum(group_totals.values(), Decimal("0"))
    ws.append(["ИТОГО", _money(grand)])
    ws.cell(row=ws.max_row, column=2).number_format = MONEY
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    _autosize(ws)

    for group in tree:
        sheet = wb.create_sheet(_sheet_title(group["name"], used_titles))
        sheet.append(["Статья расходов", "Сумма"])
        _style_header(sheet, 1, 2)
        for item in group["items"]:
            sheet.append([item, _money(amounts.get((group["name"], item), Decimal("0")))])
            sheet.cell(row=sheet.max_row, column=2).number_format = MONEY
        _autosize(sheet)

    filename = f"expenses_{d_from.isoformat()}_{d_to.isoformat()}.xlsx"
    return _wb_to_payload(wb, filename)


async def _report_attendance(session: AsyncSession, d_from: date, d_to: date) -> dict[str, Any]:
    employees = list(
        (
            await session.execute(
                select(Employee)
                .options(selectinload(Employee.position))
                .where(Employee.is_active.is_(True))
                .order_by(Employee.full_name)
            )
        )
        .scalars()
        .all()
    )
    shifts = await _shifts_in_range(session, d_from, d_to)
    shift_by_date = {s.shift_date: s for s in shifts}
    shift_ids = [s.id for s in shifts]

    events: list[AttendanceEvent] = []
    absences: list[Absence] = []
    if shift_ids:
        events = list(
            (
                await session.execute(
                    select(AttendanceEvent)
                    .where(AttendanceEvent.shift_id.in_(shift_ids))
                    .order_by(AttendanceEvent.id)
                )
            )
            .scalars()
            .all()
        )
        absences = list(
            (await session.execute(select(Absence).where(Absence.shift_id.in_(shift_ids)))).scalars().all()
        )

    events_map: dict[tuple[int, int], list[AttendanceEvent]] = defaultdict(list)
    for ev in events:
        events_map[(ev.shift_id, ev.employee_id)].append(ev)
    absence_map: dict[tuple[int, int], Absence] = {
        (a.shift_id, a.employee_id): a for a in absences
    }

    days: list[date] = []
    cur = d_from
    while cur <= d_to:
        days.append(cur)
        cur += timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "ПОСЕЩАЕМОСТЬ"
    headers = ["Сотрудник", "Должность", *[d.strftime("%d.%m.%Y") for d in days], "План", "Факт"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for emp in employees:
        pos = emp.position
        pos_name = pos.name if pos else ""
        plan_day = 0
        if pos and not pos.no_schedule and pos.work_start_time and pos.work_end_time:
            plan_day = _minutes_between(pos.work_start_time, pos.work_end_time)

        row: list[Any] = [emp.full_name, pos_name]
        plan_total = 0
        fact_total = 0
        for day in days:
            shift = shift_by_date.get(day)
            if not shift:
                row.append("")
                continue
            absence = absence_map.get((shift.id, emp.id))
            if absence:
                row.append(ABSENCE_LABELS.get(absence.kind, absence.kind))
                continue
            day_events = events_map.get((shift.id, emp.id), [])
            intervals: list[str] = []
            pending_arrival = None
            day_fact = 0
            for ev in day_events:
                if ev.event_type == "arrival":
                    pending_arrival = ev.event_time
                elif ev.event_type == "departure" and pending_arrival is not None:
                    mins = _minutes_between(pending_arrival, ev.event_time)
                    day_fact += mins
                    intervals.append(
                        f"{pending_arrival.strftime('%H:%M')}–{ev.event_time.strftime('%H:%M')}"
                    )
                    pending_arrival = None
            row.append(", ".join(intervals))
            fact_total += day_fact
            if plan_day:
                plan_total += plan_day

        row.append(_format_duration(plan_total) if plan_total else "")
        row.append(_format_duration(fact_total) if fact_total else "")
        ws.append(row)

    _autosize(ws)
    filename = f"attendance_{d_from.isoformat()}_{d_to.isoformat()}.xlsx"
    return _wb_to_payload(wb, filename)


def _describe_log(log: ActionLog) -> str:
    action = log.action or ""
    details = (log.details or "").strip()
    entity = (log.entity or "").strip()
    eid = log.entity_id

    mapping = {
        "auth.login": f"Вход в систему: {log.user_login or details}",
        "cash.update": f"Обновление кассы: {details}",
        "cash.update.forced_mismatch": f"Касса сохранена с расхождением: {details}",
        "money.income": f"Движение ДС (приход): {details}",
        "money.expense": f"Движение ДС (расход): {details}",
        "money.update": f"Изменение денежной операции: {details}",
        "money.delete": f"Удаление денежной операции: {details}",
        "attendance.arrival": f"Посещаемость (отметка приход): сотрудник {eid}, время {details}",
        "attendance.departure": f"Посещаемость (отметка уход): сотрудник {eid}, время {details}",
        "attendance.clear": f"Очистка посещаемости: сотрудник {eid}",
        "absence.set": f"отсутствие: сотрудник {eid}, тип {details}",
        "absence.remove": f"удаление отсутствия: сотрудник {eid}, {details}",
        "shift.delete": f"Удаление смены: {details}",
        "admin.user.create": f"Создан пользователь: {details}",
        "admin.position.create": f"Должность создана: {details}",
        "admin.position.archive": f"Должность архивирована: {details}",
        "admin.position.restore": f"Должность восстановлена: {details}",
        "admin.employee.create": f"Сотрудник: {details}",
        "admin.vault.create": f"Денежное хранилище: {details}",
        "accountant.income.cash": f"Бухгалтерия, приход (кассовый): {details}",
        "accountant.expense.cash": f"Бухгалтерия, расход (кассовый): {details}",
        "accountant.income.accrual": f"Бухгалтерия, приход (начисление): {details}",
        "accountant.expense.accrual": f"Бухгалтерия, расход (начисление): {details}",
        "accountant.operation.delete": f"Удаление бухгалтерской операции: {details}",
        "accountant.debt.delete": f"Удаление задолженности: {details}",
    }
    if action in mapping:
        return mapping[action]
    if details:
        return f"{action}: {details}" if action else details
    if entity and eid is not None:
        return f"{action} ({entity} #{eid})"
    return action or "Событие"


def _format_log_datetime(dt: datetime | None) -> str:
    if not dt:
        return ""
    local = dt.astimezone(app_tz()) if dt.tzinfo else dt.replace(tzinfo=app_tz())
    ms = int(local.microsecond / 1000)
    return local.strftime("%Y-%m-%d %H:%M:%S") + f":{ms:03d}"


async def _report_logs(session: AsyncSession, d_from: date, d_to: date) -> dict[str, Any]:
    start_dt = datetime.combine(d_from, datetime.min.time(), tzinfo=app_tz())
    end_dt = datetime.combine(d_to, datetime.max.time(), tzinfo=app_tz())
    logs = list(
        (
            await session.execute(
                select(ActionLog)
                .where(ActionLog.created_at >= start_dt, ActionLog.created_at <= end_dt)
                .order_by(ActionLog.created_at.desc())
            )
        )
        .scalars()
        .all()
    )
    users = {
        u.id: u
        for u in (await session.execute(select(User))).scalars().all()
    }
    users_by_login = {u.login: u for u in users.values()}

    wb = Workbook()
    ws = wb.active
    ws.title = "LOGS"
    ws.append(["Логин пользователя", "Тип пользователя", "Описание события", "ДАТА И ВРЕМЯ"])
    _style_header(ws, 1, 4)
    for log in logs:
        user = users.get(log.user_id) if log.user_id else None
        if user is None and log.user_login:
            user = users_by_login.get(log.user_login)
        role = user.role if user else ""
        ws.append(
            [
                log.user_login or (user.login if user else ""),
                role,
                _describe_log(log),
                _format_log_datetime(log.created_at),
            ]
        )
    _autosize(ws)
    filename = f"logs_{d_from.isoformat()}_{d_to.isoformat()}.xlsx"
    return _wb_to_payload(wb, filename)
